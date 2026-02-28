#!/usr/bin/env python3
"""
Export Google Sheet data to a formatted Excel file for Power BI.

Reads load out data and inventory from Google Sheet, calculates dashboard
metrics (Load Out count, Ready = Total - Redress - LoadOut), and exports
to a styled Excel workbook.

Usage:
    python export_dashboard.py              # Export to output/dashboard.xlsx
    python export_dashboard.py --open       # Export and open in Power BI
    python export_dashboard.py --sync       # Also sync new tools to Inventory
"""

import os
import sys
import argparse
from collections import Counter
from pathlib import Path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DEFAULT_SHEET_ID = "1wK92FpXq-07LdYYPCwZi7-C2vruLPs59JM14w4nAggs"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
PROJECT_DIR = Path(__file__).resolve().parent.parent.parent
OUTPUT_PATH = PROJECT_DIR / "output" / "dashboard.xlsx"


def get_credentials():
    """Load OAuth credentials."""
    for token_path in ["token_gmail.json", "token.json"]:
        full_path = PROJECT_DIR / token_path
        if full_path.exists():
            creds = Credentials.from_authorized_user_file(str(full_path), SCOPES)
            if creds and creds.valid:
                return creds
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
                with open(full_path, "w") as f:
                    f.write(creds.to_json())
                return creds
    print("Error: no valid token found.", file=sys.stderr)
    sys.exit(1)


def read_sheet_data(sheet_id):
    """Read Out and Inventory data from Google Sheet."""
    creds = get_credentials()
    service = build("sheets", "v4", credentials=creds)
    sheets = service.spreadsheets()

    # Read load out data
    out_result = sheets.values().get(
        spreadsheetId=sheet_id, range="Out!A:I"
    ).execute()
    out_rows = out_result.get("values", [])

    # Read inventory
    inv_result = sheets.values().get(
        spreadsheetId=sheet_id, range="Inventory!A:C"
    ).execute()
    inv_rows = inv_result.get("values", [])

    return out_rows, inv_rows


def sync_new_tools(sheet_id, out_rows, inv_rows):
    """Add any new descriptions from Out to Inventory in Google Sheet."""
    # Get existing tool names
    existing_tools = set()
    for row in inv_rows[1:]:
        if row:
            existing_tools.add(row[0])

    # Get all descriptions from Out
    all_descriptions = set()
    for row in out_rows[1:]:
        if len(row) >= 6 and row[5]:
            all_descriptions.add(row[5])

    new_tools = sorted(all_descriptions - existing_tools)
    if not new_tools:
        return inv_rows

    # Append to Google Sheet
    creds = get_credentials()
    service = build("sheets", "v4", credentials=creds)
    new_rows = [[t, 0, 0] for t in new_tools]
    service.spreadsheets().values().append(
        spreadsheetId=sheet_id,
        range="Inventory!A1",
        valueInputOption="RAW",
        insertDataOption="INSERT_ROWS",
        body={"values": new_rows},
    ).execute()
    print(f"  Synced {len(new_tools)} new tool(s) to Inventory")

    # Add to local inventory data
    for t in new_tools:
        inv_rows.append([t, "0", "0"])

    return inv_rows


def build_dashboard(out_rows, inv_rows):
    """Calculate dashboard metrics."""
    # Count load outs per description
    description_counts = Counter()
    for row in out_rows[1:]:
        if len(row) >= 6 and row[5]:
            description_counts[row[5]] += 1

    dashboard = []
    for row in inv_rows[1:]:
        if not row:
            continue
        tool = row[0]
        total_stock = int(row[1]) if len(row) > 1 and row[1] else 0
        redress = int(row[2]) if len(row) > 2 and row[2] else 0
        load_out = description_counts.get(tool, 0)
        ready = total_stock - redress - load_out

        dashboard.append({
            "tool": tool,
            "redress": redress,
            "ready": ready,
            "load_out": load_out,
            "total_stock": total_stock,
        })

    return dashboard


def create_excel(dashboard, out_rows, inv_rows, output_path):
    """Create formatted Excel workbook."""
    wb = Workbook()

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="334D66", end_color="334D66", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    alt_fill = PatternFill(start_color="ECF0F4", end_color="ECF0F4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="D0D0D0"),
    )

    def style_header(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def style_data_rows(ws, num_rows, num_cols, center_cols=None):
        for row_idx in range(2, num_rows + 2):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                if center_cols and col in center_cols:
                    cell.alignment = center
            if row_idx % 2 == 0:
                for col in range(1, num_cols + 1):
                    ws.cell(row=row_idx, column=col).fill = alt_fill

    # ===== Sheet 1: Dashboard =====
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_dash.sheet_properties.tabColor = "334D66"

    dash_headers = ["Tool", "Redress", "Ready", "Load Out", "Total Stock"]
    ws_dash.append(dash_headers)

    for item in dashboard:
        ws_dash.append([
            item["tool"],
            item["redress"],
            item["ready"],
            item["load_out"],
            item["total_stock"],
        ])

    style_header(ws_dash, 5)
    style_data_rows(ws_dash, len(dashboard), 5, center_cols={2, 3, 4, 5})

    # Column widths
    ws_dash.column_dimensions["A"].width = 45
    for col_letter in ["B", "C", "D", "E"]:
        ws_dash.column_dimensions[col_letter].width = 14

    # Freeze header
    ws_dash.freeze_panes = "A2"

    # ===== Sheet 2: LoadOutData =====
    ws_data = wb.create_sheet("LoadOutData")
    ws_data.sheet_properties.tabColor = "4472C4"

    for row in out_rows:
        # Pad short rows
        padded = row + [""] * (9 - len(row))
        ws_data.append(padded[:9])

    if out_rows:
        style_header(ws_data, 9)
        style_data_rows(ws_data, len(out_rows) - 1, 9)

        # Column widths
        data_widths = [50, 25, 30, 15, 14, 40, 18, 6, 22]
        for idx, w in enumerate(data_widths):
            ws_data.column_dimensions[get_column_letter(idx + 1)].width = w

        ws_data.freeze_panes = "A2"

    # ===== Sheet 3: Inventory =====
    ws_inv = wb.create_sheet("Inventory")
    ws_inv.sheet_properties.tabColor = "70AD47"

    for row in inv_rows:
        padded = row + [""] * (3 - len(row))
        ws_inv.append(padded[:3])

    if inv_rows:
        style_header(ws_inv, 3)
        style_data_rows(ws_inv, len(inv_rows) - 1, 3, center_cols={2, 3})

        ws_inv.column_dimensions["A"].width = 45
        ws_inv.column_dimensions["B"].width = 14
        ws_inv.column_dimensions["C"].width = 14

        ws_inv.freeze_panes = "A2"

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"  Exported to: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Export dashboard to Excel for Power BI")
    parser.add_argument("--sheet-id", default=DEFAULT_SHEET_ID)
    parser.add_argument("--output", default=str(OUTPUT_PATH), help="Output Excel path")
    parser.add_argument("--open", action="store_true", help="Open in Power BI after export")
    parser.add_argument("--sync", action="store_true", help="Sync new tools to Inventory sheet")
    args = parser.parse_args()

    output_path = Path(args.output)

    print("Reading Google Sheet data...")
    out_rows, inv_rows = read_sheet_data(args.sheet_id)
    print(f"  Out: {len(out_rows) - 1} load out rows")
    print(f"  Inventory: {len(inv_rows) - 1} tools")

    if args.sync:
        inv_rows = sync_new_tools(args.sheet_id, out_rows, inv_rows)

    print("Building dashboard...")
    dashboard = build_dashboard(out_rows, inv_rows)

    print("Creating Excel file...")
    create_excel(dashboard, out_rows, inv_rows, output_path)

    if args.open:
        import subprocess
        print("Opening in default application...")
        subprocess.Popen(["cmd.exe", "/c", "start", "", str(output_path)], shell=False)

    print("Done!")


if __name__ == "__main__":
    main()
